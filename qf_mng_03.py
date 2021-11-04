import calendar
import datetime
from dateutil.relativedelta import relativedelta
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from parent_class import Mod
from pandas import ExcelWriter
from openpyxl import Workbook, load_workbook


# Mng03 = Mod(ss_id=5, year=2021, month=7)
class Mng03(Mod):
    def __init__(self, ss_id, year, month, output_absolute_file_path):
        super().__init__(ss_id, year, month)

        self.fy_start_year = year
        if month <= 6:
            self.fy_start_year = year - 1

        present_month_datetime = datetime.datetime(year=year, month=month, day=calendar.monthrange(year, month)[1])
        present_month_str = f"{present_month_datetime.year}-{present_month_datetime.month}"

        last_month_datetime = present_month_datetime - relativedelta(months=1)
        last_month_str = f"{last_month_datetime.year}-{last_month_datetime.month}"

        last_fy_last_month_str = f"{self.fy_start_year}-{6}"
        """"*** pf ***"""

        def kwh_kvarh(is_kwh):
            query_string = f"""
            SELECT e.date_time, tr.name, e.value, se.id, e.is_export, e.is_import, e.is_kwh, e.is_kvarh, tr.is_auxiliary,
            se.is_transformer_low
            FROM substation_equipment se
            JOIN transformer tr ON se.transformer_id = tr.id
            JOIN energy e ON se.id = e.sub_equip_id
            WHERE (e.date_time BETWEEN '{str(self.fy_start_year)}-06-01' AND '{self.date_time}')
            AND tr.substation_id = '{self.ss_id}' 
            AND e.is_kwh={is_kwh} AND tr.is_auxiliary = 0 AND se.is_transformer_low = 1
            ORDER BY e.date_time, tr.name
            """
            return query_string

        kwh_df = pd.read_sql_query(kwh_kvarh(is_kwh=1), self.sql_db).set_index('date_time')
        kvarh_df = pd.read_sql_query(kwh_kvarh(is_kwh=0), self.sql_db).set_index('date_time')

        """calculating pf of the month"""
        total_kwh_this_month = kwh_df.loc[present_month_str]["value"].sum() - kwh_df.loc[last_month_str]["value"].sum()
        total_kvarh_this_month = kvarh_df.loc[present_month_str]["value"].sum() - kvarh_df.loc[last_month_str][
            "value"].sum()

        total_kvah_this_month = (total_kwh_this_month ** 2 + total_kvarh_this_month ** 2) ** .5
        if total_kvah_this_month != 0.0:
            self.pf = total_kwh_this_month / total_kvah_this_month
        else:
            self.pf = 'total_kvah_this_month = 0'

        """calculating cumulative pf of the Financial Year (fy)"""
        total_cumulative_kwh = kwh_df.loc[present_month_str]["value"].sum() - kwh_df.loc[last_fy_last_month_str][
            "value"].sum()
        total_cumulative_kvarh = kvarh_df.loc[present_month_str]["value"].sum() - kwh_df.loc[last_fy_last_month_str][
            "value"].sum()

        total_cumulative_kvah = (total_cumulative_kwh ** 2 + total_cumulative_kvarh ** 2) ** .5
        if total_cumulative_kvah != 0.0:
            self.cumulative_pf = total_cumulative_kwh / total_cumulative_kvah
        else:
            self.cumulative_pf = 'total_cumulative_kvah = 0'

        """***Loss***"""

        def export_import(is_export):
            query_string = f"""
                SELECT e.date_time, se.name , e.value, se.id
                FROM energy e
                JOIN substation_equipment se ON e.sub_equip_id = se.id
                WHERE (e.date_time BETWEEN '{str(self.fy_start_year)}-06-01' AND '{self.date_time}') 
                AND se.substation_id = '{self.ss_id}' 
                AND e.is_export = {is_export}
                AND (se.is_line = 1 OR se.is_feeder = 1) AND e.is_kwh = 1
                ORDER BY e.date_time, se.name
                """
            return query_string

        export_df = pd.read_sql_query(export_import(is_export=True), self.sql_db).set_index('date_time')
        import_df = pd.read_sql_query(export_import(is_export=False), self.sql_db).set_index('date_time')

        """calculating loss of the month"""
        total_export_this_month = export_df.loc[present_month_str]["value"].sum() - export_df.loc[last_month_str][
            "value"].sum()
        total_import_this_month = import_df.loc[present_month_str]["value"].sum() - import_df.loc[last_month_str][
            "value"].sum()
        if total_import_this_month >= 0.0:
            self.loss = (total_import_this_month - total_export_this_month) / total_import_this_month
        else:
            self.loss = 'total_import_this_month < 0kWh'

        """calculating cumulative loss"""
        total_cumulative_export = export_df.loc[present_month_str]["value"].sum() - \
                                  export_df.loc[last_fy_last_month_str][
                                      "value"].sum()
        total_cumulative_import = import_df.loc[present_month_str]["value"].sum() - \
                                  import_df.loc[last_fy_last_month_str][
                                      "value"].sum()
        if total_cumulative_import >= 0.0:
            self.cumulative_loss = (total_cumulative_import - total_cumulative_export) / total_cumulative_import
        else:
            self.cumulative_loss = 'total_cumulative_import < 0kWh'

        df_list = [kwh_df, kvarh_df, export_df, import_df]
        sheet_name_list = ['Tr LT kWh', 'Tr LT kVARh', 'SS Export', 'SS Import']

        wb = load_workbook('qf_mng_03_template.xlsx')

        for df, sheet_name in zip(df_list, sheet_name_list):
            # df.fillna(0)
            ws = wb.create_sheet(sheet_name)
            """copy df to sheets"""
            rows = dataframe_to_rows(df)
            for r_idx, row in enumerate(rows, 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)

        """********************************write data in main sheet********************************************"""
        ws = wb['QF_MNG_03']
        ws['F12'] = self.loss
        ws['G12'] = self.cumulative_loss
        ws['F13'] = self.pf
        ws['G13'] = self.cumulative_loss

        ws['C8'] = present_month_datetime
        ws['H8'] = datetime.datetime.now()

        ws['C5'] = f"{self.fy_start_year}-{self.fy_start_year + 1}"
        ws['C6'] = self.ss_name
        ws['C7'] = self.gmd_name
        # percent_cells = ['F12', 'G12', 'F13', 'G13']
        # for cell in percent_cells:
        #     ws[cell].number_format = '0.00%'

        """***************************************************************************************************"""

        wb.save(output_absolute_file_path)
        b = 5
