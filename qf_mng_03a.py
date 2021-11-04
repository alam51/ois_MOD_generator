import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from parent_class import Mod
from pandas import ExcelWriter
from openpyxl import Workbook, load_workbook


# Mng03 = Mod(ss_id=5, year=2021, month=7)
class Mng03(Mod):
    def __init__(self, ss_id, year, month, output_absolute_file_path):
        super().__init__(ss_id, year, month)
        fy_start_year = year
        if month <= 6:
            fy_start_year = year - 1
        else:
            fy_start_year = year
        """"*** pf ***"""

        def kwh_kvarh_ex_im(is_kwh, is_export):
            query_string = f"""
            SELECT tr.name, e.date_time, e.value, se.id, e.is_export, e.is_import, e.is_kwh, e.is_kvarh, tr.is_auxiliary,
            se.is_transformer_low
            FROM substation_equipment se
            JOIN transformer tr ON se.transformer_id = tr.id
            JOIN energy e ON se.id = e.sub_equip_id
            WHERE (e.date_time BETWEEN '{str(fy_start_year)}-06-01' AND '{self.date_time}')
            AND tr.substation_id = '{self.ss_id}' AND e.is_export = '{is_export}'
            AND e.is_kwh={is_kwh} AND tr.is_auxiliary = 0 AND se.is_transformer_low = 1
            ORDER BY tr.name, e.date_time
            """
            return query_string

        kwh_ex_df = pd.read_sql_query(kwh_kvarh_ex_im(is_kwh=1, is_export=1), self.sql_db).set_index(
            ['name', 'date_time'])
        kwh_ex_df_processed = meter_reading_to_monthly(kwh_ex_df)
        # kwh_ex_df_processed.loc[('Total', 'Monthly Energy'), 'monthly_energy'] = 5
        kwh_im_df = pd.read_sql_query(kwh_kvarh_ex_im(is_kwh=1, is_export=0), self.sql_db).set_index(
            ['name', 'date_time'])
        kwh_im_df_processed = meter_reading_to_monthly(kwh_im_df)

        kvar_ex_df = pd.read_sql_query(kwh_kvarh_ex_im(is_kwh=0, is_export=1), self.sql_db).set_index(
            ['name', 'date_time'])
        kvarh_ex_df_processed = meter_reading_to_monthly(kvar_ex_df)
        kvarh_im_df = pd.read_sql_query(kwh_kvarh_ex_im(is_kwh=0, is_export=0), self.sql_db).set_index(
            ['name', 'date_time'])
        kvarh_im_df_processed = meter_reading_to_monthly(kvarh_im_df)

        self.cumulative_kwh = kwh_ex_df_processed['monthly_energy'].sum() + kwh_im_df_processed['monthly_energy'].sum()
        self.cumulative_kvarh = kvarh_ex_df_processed['monthly_energy'].sum() + kvarh_im_df_processed['monthly_energy'].sum()
        self.cumulative_kvah = (self.cumulative_kwh ** 2 + self.cumulative_kvarh ** 2) ** .5
        self.cumulative_pf = self.cumulative_kwh / self.cumulative_kvah
        """***Loss***"""

        def export_import(is_export):
            query_string = f"""
                SELECT e.date_time, se.name , e.value, se.id
                FROM energy e
                JOIN substation_equipment se ON e.sub_equip_id = se.id
                WHERE (e.date_time BETWEEN '{str(fy_start_year)}-06-01' AND '{self.date_time}') 
                AND se.substation_id = '{self.ss_id}' 
                AND e.is_export = {is_export}
                AND (se.is_line = 1 OR se.is_feeder = 1) AND e.is_kwh = 1
                ORDER BY se.name, e.date_time
                """
            return query_string

        export_df = pd.read_sql_query(export_import(is_export=True), self.sql_db).set_index(['name', 'date_time'])
        import_df = pd.read_sql_query(export_import(is_export=False), self.sql_db).set_index(['name', 'date_time'])
        import_df_processed = meter_reading_to_monthly(import_df)
        export_df_processed = meter_reading_to_monthly(export_df)

        self.energy_export = export_df['value'].sum()
        self.energy_import = import_df['value'].sum()
        self.percent_loss = (self.energy_import - self.energy_export) / self.energy_import

        df_list = [kwh_ex_df_processed, kwh_im_df_processed, kvarh_ex_df_processed, kvarh_im_df_processed,
                   export_df_processed, import_df_processed]

        sheet_name_list = ['Tr LT kWh Export', 'Tr LT kWh Export', 'Tr LT kVARh Export', 'Tr LT kVARh Import',
                           'SS Export', 'SS Import']

        wb_template = load_workbook('qf_mng_03_template.xlsx')

        for df, sheet_name in zip(df_list, sheet_name_list):
            # df.fillna(0)
            ws = wb_template.create_sheet(sheet_name)
            """copy df to sheets"""
            rows = dataframe_to_rows(df)
            for r_idx, row in enumerate(rows, 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)

            df_reset = df.reset_index()
            df_new = df_reset.set_index('date_time')

        """********************************write data in main sheet********************************************"""
        ws = wb_template['QF_MNG_03']
        ws['G13'] = self.cumulative_pf

        """***************************************************************************************************"""

        wb_template.save('qf_mng_03.xlsx')
        b = 5


def meter_reading_to_monthly(meter_reading_df):
    df = meter_reading_df.copy()
    running_name = ''
    running_name_changed = False
    # for serial, [name, date_time] in enumerate(meter_reading_df.index):
    try:
        for serial, [name, date_time] in enumerate(meter_reading_df.index):
            if name != running_name:
                running_name = name
                previous_date_time = date_time
                if date_time.month != 6:
                    df.loc[(name, date_time), 0] = 'Could not read: Energy reading of June, Previous year'
                    return df

            else:
                df.loc[(name, date_time), 'monthly_energy'] = meter_reading_df.loc[(name, date_time), 'value'] - \
                                                              meter_reading_df.loc[(name,
                                                                                    previous_date_time), 'value']
                previous_date_time = date_time

        return df

    except:
        pass
